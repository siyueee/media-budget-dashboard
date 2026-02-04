import logging
import sys
import pandas as pd
import json
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from typing import Tuple, Union, Dict, List

# ==============================================================================
#  全局配置
# ==============================================================================
CONFIG = {
    'INPUT_FILE': 'data/OCPX业务报表 (4).xlsx',
    'ORDER_FILE': r'订单明细.xlsx',
    'OUTPUT_FILE': r'媒体预算日数据.xlsx',
    'KPI_CONFIG_FILE': r'订单考核.json',
    'RAW_DATA_SHEET': '广告调度交叉数据去重报表',
    'ORDER_SHEET': '订单明细',
    'PLATFORM_FILTERS': ['45_微博', '120_网易新闻', '130_优酷(媒体)CPA', '103_喜马拉雅-非官预算', '23_网易有道7.0','144_合攒'],
    'REQUIRED_COLUMNS_MAPPING': {
        '广告主激活量': '激活量',
        '上报广告主次数': '上报广告主次数',  # 正常提取
        '上报广告主曝光数': '上报广告主曝光数',  # 正常提取
        '次日回访量': '次日回访量',
        '2日留存数': '2日留存数',
        '下单量': '下单量',
        '新登量': '新登量',
        '付费数': '付费数',
        '首购量': '首购量',
        '唤醒量': '唤醒量'
    },
    'ORDER_COLUMNS_TO_MERGE': ['合作价格', '需求量级', '考核备注', '产品', '渠道号', '回传维度', '考核', '考核数值'],
}


class DataProcessor:
    def __init__(self, config: Dict):
        self.config = config
        self.logger = self._setup_logging()
        self.raw_df = None
        self.order_df = None
        self.kpi_map = self._load_kpi_config()

    def _setup_logging(self):
        logger = logging.getLogger('DataProcessor')
        logger.setLevel(logging.INFO)
        if not logger.handlers:
            handler = logging.StreamHandler(sys.stdout)
            handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
            logger.addHandler(handler)
        return logger

    def _load_kpi_config(self):
        try:
            with open(self.config['KPI_CONFIG_FILE'], 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}

    def read_all_inputs(self):
        self.logger.info("🚀 正在读取输入文件...")
        self.raw_df = pd.read_excel(self.config['INPUT_FILE'], sheet_name=self.config['RAW_DATA_SHEET'])
        self.raw_df.columns = self.raw_df.columns.str.strip()
        if '广告主激活量' in self.raw_df.columns:
            self.raw_df.rename(columns={'广告主激活量': '激活量'}, inplace=True)
        self.raw_df['日期'] = pd.to_datetime(self.raw_df['日期'])
        self.order_df = pd.read_excel(self.config['ORDER_FILE'], sheet_name=self.config['ORDER_SHEET'])
        self.order_df.columns = self.order_df.columns.str.strip()
        self.order_df['配置号'] = self.order_df['配置号'].astype(str)
        self.order_df.drop_duplicates(subset=['配置号'], inplace=True)
        if '广告主平台配置名称' in self.raw_df.columns:
            self.raw_df['配置号'] = self.raw_df['广告主平台配置名称'].str.split('_', n=1).str.get(1).astype(str)

    def split_data_by_platform(self) -> Dict[str, pd.DataFrame]:
        platform_dfs = {}
        for platform in self.config['PLATFORM_FILTERS']:
            p_esc = platform.replace('(', r'\(').replace(')', r'\)')
            mask = self.raw_df['广告主平台名称'].str.contains(p_esc, na=False)
            platform_dfs[platform] = self.raw_df[mask].copy()
        return platform_dfs

    def _create_pivot_table(self, df: pd.DataFrame, is_detail: bool = False) -> pd.DataFrame:
        idx = ['日期', '配置号']
        if is_detail:
            for col in ['媒体平台名称', '调度中心ID']:
                if col in df.columns: idx.append(col)
        agg_targets = list(self.config['REQUIRED_COLUMNS_MAPPING'].values())
        actual_agg = {col: 'sum' for col in agg_targets if col in df.columns}
        pivot_df = df.pivot_table(index=idx, values=list(actual_agg.keys()), aggfunc='sum').reset_index()
        for col in agg_targets:
            if col not in pivot_df.columns: pivot_df[col] = 0
        pivot_df = pivot_df.sort_values(by=idx, ascending=True)
        group_keys = ['配置号']
        if is_detail:
            if '媒体平台名称' in pivot_df.columns: group_keys.append('媒体平台名称')
            if '调度中心ID' in pivot_df.columns: group_keys.append('调度中心ID')
        pivot_df['次日回访_平移'] = pivot_df.groupby(group_keys)['次日回访量'].shift(-1)
        pivot_df['2日留存_平移'] = pivot_df.groupby(group_keys)['2日留存数'].shift(-1)
        if is_detail:
            pivot_df.rename(columns={'媒体平台名称': '媒体平台', '调度中心ID': '调度中心id'}, inplace=True)
        return pivot_df

    def _calculate_kpi(self, df: pd.DataFrame, platform_name: str) -> pd.DataFrame:
        def get_kpi_info(row):
            order_kpi = str(row.get('考核', '')).strip()
            if order_kpi and order_kpi != 'nan' and order_kpi != '':
                if '次留' in order_kpi: return '次留率'
                if '下单' in order_kpi: return '下单率'
                if '付费' in order_kpi: return '付费率'
                if '首购' in order_kpi: return '首购率'
                return order_kpi
            return self.kpi_map.get('配置号映射', {}).get(row['配置号'], {}).get('考核项', None)

        df['考核项'] = df.apply(get_kpi_info, axis=1)

        def calc_res(row):
            term = row['考核项']
            if not term: return None
            act = row.get('激活量', 0)
            if act == 0: return 0
            res = 0
            if term == '次留率':
                res = (row.get('2日留存_平移', 0) if "103_喜马拉雅" in platform_name else row.get('次日回访_平移',
                                                                                                  0)) / act
            elif term == '下单率':
                res = row.get('下单量', 0) / act
            elif term == '付费率':
                res = row.get('付费数', 0) / act
            elif term == '首购率':
                res = row.get('首购量', 0) / act
            return round(res, 3) if res else 0

        df['考核结果'] = df.apply(calc_res, axis=1)
        return df

    def _adjust_column_order(self, df: pd.DataFrame, platform_name: str, is_detail: bool = False) -> pd.DataFrame:
        if is_detail:
            cols = ['日期', '甲方', '配置号', '媒体平台', '调度中心id', '产品', '渠道号', '回传维度', '合作价格']
        else:
            cols = ['日期', '甲方', '配置号', '产品', '渠道号', '回传维度', '合作价格', '需求量级']

        # 曝光和上报插入在前面，但不参与过滤
        cols.extend(['上报广告主曝光数', '上报广告主次数', '激活量', '唤醒量'])

        if "23_网易有道" in platform_name:
            ins_pt = '唤醒量' if '唤醒量' in cols else '激活量'
            cols.insert(cols.index(ins_pt) + 1, '首购量')
        cols.extend(['下单量', '新登量', '付费数', '次日回访量'])
        if "103_喜马拉雅" in platform_name: cols.append('2日留存数')
        cols.extend(['考核项', '考核结果', '考核数值', '考核备注'])
        return df[[c for c in cols if c in df.columns]]

    def process_platform_data(self, df: pd.DataFrame, platform_name: str, is_detail: bool = False) -> pd.DataFrame:
        p_df = self._create_pivot_table(df, is_detail)
        m_df = pd.merge(p_df, self.order_df[['配置号'] + self.config['ORDER_COLUMNS_TO_MERGE']], on='配置号',
                        how='left')
        m_df['甲方'] = platform_name
        k_df = self._calculate_kpi(m_df, platform_name)
        final_df = self._adjust_column_order(k_df, platform_name, is_detail)
        if '日期' in final_df.columns: final_df = final_df.sort_values(by=['日期'], ascending=[True])
        return final_df

    def save_to_excel(self, data_dict: Dict[str, pd.DataFrame], output_path: str):
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sn, df in data_dict.items():
                if df.empty: continue
                chk = [c for c in ['激活量', '下单量', '付费数', '唤醒量', '首购量'] if c in df.columns]
                df = df[df[chk].sum(axis=1) > 0].copy()

                if df.empty: continue
                df['日期'] = df['日期'].dt.strftime('%Y-%m-%d')
                sn_safe = sn.replace('(', '').replace(')', '')[:31]
                df.to_excel(writer, sheet_name=sn_safe, index=False)
                self._format_excel(writer.sheets[sn_safe], df)

    def _format_excel(self, ws, df):
        left_align = Alignment(horizontal='left', vertical='center')
        for col in ws.columns:
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = 17
            for cell in col: cell.alignment = left_align
        res_idx = next((i + 1 for i, c in enumerate(df.columns) if c == '考核结果'), None)
        std_idx = next((i + 1 for i, c in enumerate(df.columns) if c == '考核数值'), None)
        for row in range(2, ws.max_row + 1):
            if res_idx:
                c = ws.cell(row=row, column=res_idx)
                if c.value is not None: c.number_format = '0.0%'
            if std_idx:
                c = ws.cell(row=row, column=std_idx)
                if c.value is not None:
                    try:
                        v = str(c.value)
                        if '%' in v: c.value = float(v.replace('%', '')) / 100
                        c.number_format = '0.00%'
                    except:
                        pass
            if res_idx and std_idx:
                try:
                    rv = float(ws.cell(row=row, column=res_idx).value or 0)
                    sv = str(ws.cell(row=row, column=std_idx).value or "0")
                    sn = float(sv.replace('%', '')) / 100 if '%' in sv else float(sv)
                    if rv < sn: ws.cell(row=row, column=res_idx).font = Font(color="FF0000", bold=True)
                except:
                    pass


def main():
    proc = DataProcessor(CONFIG)
    try:
        proc.read_all_inputs()
        raw_dfs = proc.split_data_by_platform()

        # 1. 处理并保存主表（保持原样，不加“归属”）
        res_std = {p: proc.process_platform_data(d, p, False) for p, d in raw_dfs.items()}
        proc.save_to_excel(res_std, CONFIG['OUTPUT_FILE'])

        # 2. 处理明细表
        res_det = {p: proc.process_platform_data(d, p, True) for p, d in raw_dfs.items()}

        # --- 新增逻辑：仅在明细表中加入“归属”列 ---
        if '归属' in proc.order_df.columns:
            # 创建 配置号 -> 归属 的映射字典
            mapping = proc.order_df.set_index('配置号')['归属'].to_dict()

            for p in res_det:
                if not res_det[p].empty:
                    # 在指定位置插入“归属”列（比如插在“甲方”后面，即第2列）
                    res_det[p].insert(2, '归属', res_det[p]['配置号'].map(mapping))
        # ---------------------------------------

        # 3. 保存明细表
        out_det = str(Path(CONFIG['OUTPUT_FILE']).with_name(Path(CONFIG['OUTPUT_FILE']).stem + "_附带明细.xlsx"))
        proc.save_to_excel(res_det, out_det)

        print(f"✅ 完成！\n主表已保持纯净，明细表已成功插入“归属”列。")
    except Exception as e:
        logging.critical(f"崩溃: {e}", exc_info=True)

if __name__ == "__main__":
    main()
    print("✅ 数据处理完成！")
    print("🚀 正在尝试同步到 GitHub...")